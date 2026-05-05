#!/usr/bin/env python3
"""
run_portfolio_optimizer.py  (mbot)

Lädt alle Configs, führt Backtests durch und wählt das beste Portfolio
per Portfolio-Simulation + Calmar-Greedy. Schreibt active_strategies in settings.json.

Aufruf:
  python3 run_portfolio_optimizer.py              # interaktiv
  python3 run_portfolio_optimizer.py --auto-write # automatisch (Scheduler)
"""
import os
import sys
import json
import argparse
from datetime import date, timedelta
from tqdm import tqdm

PROJECT_ROOT  = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))

CONFIGS_DIR   = os.path.join(PROJECT_ROOT, 'src', 'mbot', 'strategy', 'configs')
SETTINGS_PATH = os.path.join(PROJECT_ROOT, 'settings.json')
SECRET_PATH   = os.path.join(PROJECT_ROOT, 'secret.json')

B  = '\033[1;37m'
G  = '\033[0;32m'
Y  = '\033[1;33m'
R  = '\033[0;31m'
NC = '\033[0m'

DEFAULT_LOOKBACK_DAYS = 1095  # ~3 Jahre als Standard


def _scan_configs() -> list:
    if not os.path.isdir(CONFIGS_DIR):
        return []
    return sorted([
        os.path.join(CONFIGS_DIR, f)
        for f in os.listdir(CONFIGS_DIR)
        if f.endswith('.json')
    ])


def _make_exchange():
    from mbot.utils.exchange import Exchange
    with open(SECRET_PATH) as f:
        secrets = json.load(f)
    accounts = secrets.get('mbot', [])
    if not accounts:
        raise RuntimeError("Keine 'mbot'-Accounts in secret.json")
    return Exchange(accounts[0])


def _build_results_dict(config_files: list, risk_config: dict, capital: float,
                         exchange, start_date: str, end_date: str) -> dict:
    from mbot.analysis.backtester import load_data, run_backtest
    results_dict = {}
    for path in tqdm(config_files, desc='Lade Configs & Backtests'):
        fname = os.path.basename(path)
        try:
            with open(path) as f:
                cfg = json.load(f)
            market        = cfg.get('market', {})
            symbol        = market.get('symbol', '')
            timeframe     = market.get('timeframe', '')
            signal_config = cfg.get('signal', {})
            if not symbol or not timeframe:
                continue
            data = load_data(exchange, symbol, timeframe, start_date, end_date)
            if data is None or data.empty or len(data) < 50:
                print(f"  {Y}Uebersprungen (keine Daten): {fname}{NC}")
                continue
            result = run_backtest(data, signal_config, risk_config, capital, symbol)
            if not result or result.get('total_trades', 0) == 0:
                continue
            results_dict[fname] = {**result, 'symbol': symbol, 'timeframe': timeframe}
        except Exception as e:
            print(f"  {Y}Fehler bei {fname}: {e}{NC}")
    return results_dict


def _simulate_current_portfolio(settings: dict, results_dict: dict,
                                 start_capital: float) -> dict | None:
    """Simuliert das aktuell aktive Portfolio mit den vorhandenen Backtest-Daten."""
    from mbot.analysis.portfolio_simulator import run_portfolio_simulation
    current = [
        s for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
        if s.get('active')
    ]
    if not current:
        return None
    subset = {}
    for s in current:
        sym, tf = s.get('symbol', ''), s.get('timeframe', '')
        for fname, rd in results_dict.items():
            if rd.get('symbol') == sym and rd.get('timeframe') == tf:
                subset[fname] = rd
                break
    if not subset:
        return None
    return run_portfolio_simulation(subset, start_capital)


def _write_to_settings(selected_files: list, results_dict: dict) -> None:
    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    existing     = settings.get('live_trading_settings', {}).get('active_strategies', [])
    existing_map = {(s.get('symbol'), s.get('timeframe')): s for s in existing}
    new_strategies = []
    for fname in selected_files:
        rd        = results_dict.get(fname, {})
        symbol    = rd.get('symbol', '')
        timeframe = rd.get('timeframe', '')
        if not symbol or not timeframe:
            continue
        base  = existing_map.get((symbol, timeframe), {})
        entry = {**base, 'symbol': symbol, 'timeframe': timeframe, 'active': True}
        new_strategies.append(entry)
    lt = settings.setdefault('live_trading_settings', {})
    lt['active_strategies']          = new_strategies
    lt['use_auto_optimizer_results'] = True
    with open(SETTINGS_PATH, 'w') as f:
        json.dump(settings, f, indent=4)


BOT_NAME = 'mbot'


def _get_telegram_creds():
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json')) as f:
            s = json.load(f)
        tg = s.get('telegram', {})
        t, c = tg.get('bot_token', ''), tg.get('chat_id', '')
        return (t, c) if t and c else (None, None)
    except Exception:
        return None, None


def _send_telegram(msg):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        requests.post(f'https://api.telegram.org/bot{token}/sendMessage',
                      data={'chat_id': chat, 'text': msg}, timeout=10)
    except Exception:
        pass


def _send_telegram_doc(fpath, caption=''):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        with open(fpath, 'rb') as fh:
            requests.post(f'https://api.telegram.org/bot{token}/sendDocument',
                          data={'chat_id': chat, 'caption': caption},
                          files={'document': fh}, timeout=30)
    except Exception:
        pass


def generate_trades_excel(final, results_dict, capital, start_date, end_date):
    """Erstellt Excel-Tabelle mit allen Portfolio-Trades."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f'  {Y}openpyxl nicht installiert — Excel uebersprungen.{NC}')
        return None

    trades = final.get('trades', [])
    if not trades:
        return None

    rows = []
    for i, t in enumerate(trades, 1):
        fname  = t.get('_strategy_key', '')
        rd     = results_dict.get(fname, {})
        symbol = rd.get('symbol', '?')
        tf     = t.get('_timeframe', rd.get('timeframe', '?'))
        result = t.get('result', '')
        pnl    = t.get('portfolio_pnl_usdt', 0.0)
        equity = t.get('portfolio_capital_after', 0.0)
        rows.append({
            'Nr':            i,
            'Datum':         str(t.get('entry_time', ''))[:16].replace('T', ' '),
            'Symbol':        symbol,
            'Timeframe':     tf,
            'Richtung':      str(t.get('direction', '?')).upper(),
            'Ergebnis':      'TP erreicht' if result == 'win' else 'SL erreicht',
            'PnL (USDT)':    round(pnl, 4),
            'Gesamtkapital': round(equity, 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'
    hdr  = PatternFill('solid', fgColor='1E3A5F')
    win  = PatternFill('solid', fgColor='D6F4DC')
    loss = PatternFill('solid', fgColor='FAD7D7')
    alt  = PatternFill('solid', fgColor='F2F2F2')
    brd  = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                  top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    cw   = {'Nr': 6, 'Datum': 18, 'Symbol': 22, 'Timeframe': 12, 'Richtung': 10,
             'Ergebnis': 14, 'PnL (USDT)': 14, 'Gesamtkapital': 16}
    hdrs = list(rows[0].keys())
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = brd
        ws.column_dimensions[get_column_letter(c)].width = cw.get(h, 14)
    ws.row_dimensions[1].height = 22
    for ri, row in enumerate(rows, 2):
        f = win if row['Ergebnis'] == 'TP erreicht' else (loss if ri % 2 == 0 else alt)
        for c, key in enumerate(hdrs, 1):
            cell = ws.cell(row=ri, column=c, value=row[key])
            cell.fill = f
            cell.border = brd
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('PnL (USDT)', 'Gesamtkapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[ri].height = 18
    pnl = final.get('total_pnl_pct', 0)
    dd  = final.get('max_drawdown', 0)
    wr  = final.get('win_rate', 0)
    eq  = final.get('end_capital', rows[-1]['Gesamtkapital'] if rows else capital)
    n   = final.get('total_trades', len(trades))
    sr  = len(rows) + 3
    for label, val in [('Zeitraum', f'{start_date} -> {end_date}'), ('Trades', n),
                        ('Win-Rate', f'{wr:.1f}%'), ('PnL', f'{pnl:+.1f}%'),
                        ('Endkapital', f'{eq:.2f} USDT'), ('Max Drawdown', f'{dd:.1f}%')]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=val)
        sr += 1
    outfile = f'/tmp/{BOT_NAME}_trades.xlsx'
    wb.save(outfile)
    print(f'  {G}✓ Excel erstellt: {outfile}{NC}')
    return outfile


def generate_equity_html(final, results_dict_subset, capital, start_date, end_date, labels=None):
    """Delegiert an interactive_chart.generate_portfolio_chart (mit Einzel-Kurven + TP/SL-Marker)."""
    try:
        from mbot.analysis.interactive_chart import generate_portfolio_chart
        path = generate_portfolio_chart(results_dict_subset, final, capital, start_date, end_date)
        if path:
            print(f'  {G}✓ Chart erstellt: {path}{NC}')
        return path
    except Exception as e:
        print(f'  {Y}Chart fehlgeschlagen: {e}{NC}')
        return None


def _do_replot(settings: dict, capital: float, start_date: str, end_date: str) -> int:
    print(f"\n{'─'*72}")
    print(f"{B}  mbot — Replot (aktives Portfolio){NC}")
    print(f"  Kapital: {capital:.0f} USDT | Zeitraum: {start_date} → {end_date}")
    print(f"{'─'*72}\n")

    active = [s for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
              if s.get('active')]
    if not active:
        print(f"{R}  Keine aktiven Strategien in settings.json.{NC}")
        return 1

    active_pairs = {(s['symbol'], s['timeframe']) for s in active}
    matching = []
    for path in _scan_configs():
        try:
            with open(path) as f:
                cfg = json.load(f)
            m = cfg.get('market', {})
            if (m.get('symbol'), m.get('timeframe')) in active_pairs:
                matching.append(path)
        except Exception:
            pass

    if not matching:
        print(f"{R}  Keine Config-Dateien fuer aktive Strategien gefunden.{NC}")
        return 1

    print(f"  {len(matching)} Config(s) gefunden.\n")
    try:
        exchange = _make_exchange()
    except Exception as e:
        print(f"{R}  Exchange-Verbindung fehlgeschlagen: {e}{NC}")
        return 1

    risk_config  = settings.get('risk', {})
    results_dict = _build_results_dict(matching, risk_config, capital, exchange,
                                       start_date, end_date)
    if not results_dict:
        print(f"{R}  Keine Backtests erfolgreich.{NC}")
        return 1

    from mbot.analysis.portfolio_simulator import run_portfolio_simulation
    final = run_portfolio_simulation(results_dict, capital)

    selected_files = list(results_dict.keys())
    pnl = final.get('total_pnl_pct', 0)
    dd  = final.get('max_drawdown', 0)
    n   = final.get('total_trades', 0)
    wr  = final.get('win_rate', 0)
    eq  = final.get('end_capital', 0)

    print(f"\n{'='*72}")
    print(f"{B}  Replot — {len(selected_files)} Strategie(n){NC}\n")
    for fname in selected_files:
        rd = results_dict.get(fname, {})
        print(f"  {G}✓{NC} {rd.get('symbol', fname):<26} / {rd.get('timeframe', ''):<6}")
    print(f"\n  Endkapital: {eq:.2f} USDT  | PnL: {pnl:+.1f}%  | MaxDD: {dd:.2f}%")
    print(f"{'='*72}\n")

    summary = (f"{BOT_NAME} Replot\n"
               f"{len(selected_files)} Strategien | {n} Trades | WR: {wr:.1f}%\n"
               f"PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}% | Equity: {eq:.2f} USDT\n"
               f"Zeitraum: {start_date} -> {end_date}")
    _send_telegram(summary)
    xlsx = generate_trades_excel(final, results_dict, capital, start_date, end_date)
    if xlsx:
        _send_telegram_doc(xlsx, caption=f'{BOT_NAME} Trades | {n} Trades | WR: {wr:.1f}% | Equity: {eq:.2f} USDT')
    html = generate_equity_html(final, results_dict, capital, start_date, end_date)
    if html:
        _send_telegram_doc(html, caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}%')
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description='mbot Portfolio-Optimizer')
    parser.add_argument('--capital',    type=float, default=None)
    parser.add_argument('--max-dd',     type=float, default=30.0)
    parser.add_argument('--start-date', type=str,   default=None)
    parser.add_argument('--end-date',   type=str,   default=None)
    parser.add_argument('--auto-write', action='store_true')
    parser.add_argument('--replot',     action='store_true',
                        help='Replot fuer aktives Portfolio (keine Re-Optimierung)')
    args = parser.parse_args()

    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    opt           = settings.get('optimization_settings', {})
    risk_config   = settings.get('risk', {})
    capital       = args.capital or float(opt.get('start_capital', 100))
    max_dd        = args.max_dd
    end_date      = args.end_date   or date.today().strftime('%Y-%m-%d')
    start_date    = args.start_date or (
        date.today() - timedelta(days=DEFAULT_LOOKBACK_DAYS)
    ).strftime('%Y-%m-%d')
    max_positions = int(settings.get('live_trading_settings', {}).get('max_open_positions', 10))

    if args.replot:
        return _do_replot(settings, capital, start_date, end_date)

    print(f"\n{'─'*72}")
    print(f"{B}  mbot — Automatische Portfolio-Optimierung{NC}")
    print(f"  Portfolio-Simulation + Calmar-Greedy (MaxDD ≤ {max_dd:.0f}%)")
    print(f"  Kapital: {capital:.0f} USDT | Positionen: max {max_positions} | "
          f"Zeitraum: {start_date} → {end_date}")
    print(f"{'─'*72}\n")

    config_files = _scan_configs()
    if not config_files:
        print(f"{R}  Keine Configs in {CONFIGS_DIR}{NC}")
        print(f"  → Zuerst run_pipeline.sh ausfuehren!\n")
        return 1

    try:
        exchange = _make_exchange()
    except Exception as e:
        print(f"{R}  Exchange-Verbindung fehlgeschlagen: {e}{NC}")
        return 1

    print(f"  {len(config_files)} Config(s) gefunden.\n")
    results_dict = _build_results_dict(config_files, risk_config, capital, exchange,
                                        start_date, end_date)
    if not results_dict:
        print(f"{R}  Keine Backtests erfolgreich.{NC}")
        return 1

    from mbot.analysis.portfolio_simulator import find_best_portfolio
    portfolio = find_best_portfolio(results_dict, capital, max_dd, verbose=True)

    if not portfolio or not portfolio.get('selected'):
        print(f"{R}  Kein Portfolio erfuellt die Bedingungen (MaxDD ≤ {max_dd:.0f}%).{NC}\n")
        return 0

    selected_files = portfolio['selected'][:max_positions]
    final          = portfolio.get('portfolio') or {}

    print(f"\n{'='*72}")
    print(f"{B}  Optimales Portfolio — {len(selected_files)} Strategie(n){NC}\n")
    for fname in selected_files:
        rd = results_dict.get(fname, {})
        print(f"  {G}✓{NC} {rd.get('symbol', fname):<26} / {rd.get('timeframe', ''):<6}")
    if final:
        pnl = final.get('total_pnl_pct', 0)
        print(f"\n  Endkapital: {final.get('end_capital', 0):.2f} USDT  "
              f"| PnL: {pnl:+.1f}%  "
              f"| MaxDD: {final.get('max_drawdown', 0):.2f}%")
    print(f"{'='*72}\n")

    current_set = {
        (s.get('symbol'), s.get('timeframe'))
        for s in settings.get('live_trading_settings', {}).get('active_strategies', [])
        if s.get('active')
    }
    new_set = {
        (results_dict.get(f, {}).get('symbol'), results_dict.get(f, {}).get('timeframe'))
        for f in selected_files
    }

    cur_result  = _simulate_current_portfolio(settings, results_dict, capital)
    cur_cap     = cur_result.get('end_capital', 0) if cur_result else 0
    new_cap     = final.get('end_capital', 0)
    if cur_result:
        print(f"  Aktuelles Portfolio: {cur_cap:.2f} USDT  "
              f"| PnL: {cur_result.get('total_pnl_pct', 0):+.1f}%  "
              f"| MaxDD: {cur_result.get('max_drawdown', 0):.2f}%")
        print(f"  Neues Portfolio:     {new_cap:.2f} USDT  "
              f"| PnL: {final.get('total_pnl_pct', 0):+.1f}%  "
              f"| MaxDD: {final.get('max_drawdown', 0):.2f}%\n")

    if args.auto_write:
        if cur_result and new_cap <= cur_cap:
            print(f"{Y}  Neues Portfolio ({new_cap:.2f} USDT) nicht besser als aktuelles "
                  f"({cur_cap:.2f} USDT) — keine Aenderung.{NC}\n")
        else:
            _write_to_settings(selected_files, results_dict)
            print(f"{G}✓ settings.json aktualisiert — {len(selected_files)} Strategie(n).{NC}\n")
    else:
        if current_set == new_set:
            print(f"{Y}  Portfolio unveraendert — keine Aenderung noetig.{NC}\n")
        else:
            try:
                ans = input("  Optimales Portfolio in settings.json eintragen? (j/n): ").strip().lower()
            except (EOFError, KeyboardInterrupt):
                ans = 'n'
            if ans in ('j', 'ja', 'y', 'yes'):
                _write_to_settings(selected_files, results_dict)
                print(f"{G}✓ settings.json aktualisiert.{NC}\n")
            else:
                print(f"{Y}  settings.json NICHT geaendert.{NC}\n")

    # ── Reports & Telegram ──────────────────────────────────────────────────
    if args.auto_write:
        pnl = final.get('total_pnl_pct', 0)
        dd  = final.get('max_drawdown', 0)
        n   = final.get('total_trades', 0)
        wr  = final.get('win_rate', 0)
        eq  = final.get('end_capital', 0)
        summary = (f"{BOT_NAME} Auto-Optimizer\n"
                   f"{len(selected_files)} Strategien | {n} Trades | WR: {wr:.1f}%\n"
                   f"PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}% | Equity: {eq:.2f} USDT\n"
                   f"Zeitraum: {start_date} -> {end_date}")
        _send_telegram(summary)
        subset = {f: results_dict[f] for f in selected_files}
        xlsx = generate_trades_excel(final, results_dict, capital, start_date, end_date)
        if xlsx:
            _send_telegram_doc(xlsx, caption=f'{BOT_NAME} Trades | {n} Trades | WR: {wr:.1f}% | Equity: {eq:.2f} USDT')
        html = generate_equity_html(final, subset, capital, start_date, end_date)
        if html:
            _send_telegram_doc(html, caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}%')

    return 0


if __name__ == '__main__':
    sys.exit(main())
