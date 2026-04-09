# src/mbot/analysis/show_results.py
"""
mbot Ergebnis-Analyse

Modi:
  1) Einzel-Analyse            — jede Strategie isoliert backtesten
  2) Manuelle Portfolio-Sim    — user waehlt Strategien, gemeinsamer Kapitalpool
  3) Auto Portfolio-Optimierung — Bot findet bestes Team
  4) Interaktive Charts         — Candlestick + Trade-Marker als HTML
"""

import os
import sys
import json
import argparse
import logging
from datetime import datetime, timezone

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.append(os.path.join(PROJECT_ROOT, 'src'))

CONFIGS_DIR  = os.path.join(PROJECT_ROOT, 'src', 'mbot', 'strategy', 'configs')
RESULTS_DIR  = os.path.join(PROJECT_ROOT, 'artifacts', 'results')

GREEN  = '\033[0;32m'
BLUE   = '\033[0;34m'
YELLOW = '\033[1;33m'
RED    = '\033[0;31m'
CYAN   = '\033[0;36m'
BOLD   = '\033[1m'
NC     = '\033[0m'

logging.basicConfig(level=logging.WARNING)


# ============================================================
# Excel Export
# ============================================================

def _generate_trades_excel(portfolio: dict, start_capital: float):
    """Erstellt mbot_trades.xlsx mit allen Portfolio-Trades."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f'  {YELLOW}openpyxl nicht installiert — Excel uebersprungen. (pip install openpyxl){NC}')
        return

    trades = portfolio.get('trades', [])
    if not trades:
        print(f'  {YELLOW}Keine Trades — Excel uebersprungen.{NC}')
        return

    rows = []
    for i, t in enumerate(trades):
        symbol    = t.get('symbol', '?')
        coin      = symbol.split('/')[0] if '/' in symbol else symbol
        tf        = t.get('_timeframe', '?')
        side      = t.get('side', '?').upper()
        entry_p   = round(float(t.get('entry_price', 0)), 6)
        exit_p    = round(float(t.get('exit_price',  0)), 6)
        result    = t.get('result', '')
        ergebnis  = 'TP' if result == 'win' else 'SL'
        pnl_pct          = round(float(t.get('pnl_pct', 0)), 2)
        pnl_usdt         = round(float(t.get('portfolio_pnl_usdt', 0)), 4)
        kapital          = round(float(t.get('portfolio_capital_after', 0)), 4)
        risk_per_trade   = float(t.get('risk_per_trade_pct', 0))
        leverage         = t.get('leverage', '—')
        kapital_vor      = kapital - pnl_usdt
        einsatz          = round(kapital_vor * risk_per_trade / 100.0, 4)
        datum            = str(t.get('entry_time', ''))[:16].replace('T', ' ')

        rows.append({
            'Nr':              i + 1,
            'Datum':           datum,
            'Symbol':          coin,
            'TF':              tf,
            'Richtung':        side,
            'Hebel':           leverage,
            'Entry':           entry_p,
            'Exit':            exit_p,
            'Ergebnis':        ergebnis,
            'Einsatz (USDT)':  einsatz,
            'PnL%':            pnl_pct,
            'PnL (USDT)':      pnl_usdt,
            'Kapital':         kapital,
        })

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Trades'

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    win_fill    = PatternFill('solid', fgColor='D6F4DC')
    loss_fill   = PatternFill('solid', fgColor='FAD7D7')
    alt_fill    = PatternFill('solid', fgColor='F2F2F2')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),  right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin',  color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
    )
    col_widths = {
        'Nr': 5, 'Datum': 18, 'Symbol': 10, 'TF': 7, 'Richtung': 10,
        'Hebel': 7, 'Entry': 14, 'Exit': 14, 'Ergebnis': 9,
        'Einsatz (USDT)': 16, 'PnL%': 10, 'PnL (USDT)': 14, 'Kapital': 16,
    }

    headers = list(rows[0].keys())
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        fill = win_fill if row['Ergebnis'] == 'TP' else (loss_fill if r_idx % 2 == 0 else alt_fill)
        for col, key in enumerate(headers, 1):
            cell           = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Entry', 'Exit', 'PnL (USDT)', 'Kapital', 'Einsatz (USDT)'):
                cell.number_format = '#,##0.0000'
            elif key == 'PnL%':
                cell.number_format = '+0.00;-0.00'
        ws.row_dimensions[r_idx].height = 18

    # Zusammenfassung
    total     = len(rows)
    wins      = sum(1 for r in rows if r['Ergebnis'] == 'TP')
    pnl_total = rows[-1]['Kapital'] - start_capital if rows else 0.0
    pnl_pct_s = pnl_total / start_capital * 100 if start_capital else 0.0
    sr        = total + 3
    ws.cell(row=sr, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    for label, value in [
        ('Trades gesamt', total),
        ('Win-Rate',      f'{wins / total * 100:.1f}%' if total else '—'),
        ('PnL',           f'{pnl_pct_s:+.1f}%'),
        ('Endkapital',    f'{rows[-1]["Kapital"]:.4f} USDT' if rows else '—'),
    ]:
        ws.cell(row=sr, column=1, value=label).font = Font(bold=True)
        ws.cell(row=sr, column=2, value=value)
        sr += 1

    out_dir  = os.path.join(PROJECT_ROOT, 'artifacts', 'charts')
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'mbot_trades.xlsx')
    wb.save(out_file)
    print(f'  {GREEN}Excel gespeichert: mbot_trades.xlsx{NC}')

    # Via Telegram senden
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json')) as f:
            secrets = json.load(f)
        tg        = secrets.get('telegram', {})
        bot_token = tg.get('bot_token', '')
        chat_id   = tg.get('chat_id', '')
        if bot_token and chat_id:
            import requests
            caption = (f'mbot Trades — {total} Trades | '
                       f'WR: {wins / total * 100:.1f}% | PnL: {pnl_pct_s:+.1f}%' if total else 'mbot Trades')
            with open(out_file, 'rb') as doc:
                requests.post(
                    f'https://api.telegram.org/bot{bot_token}/sendDocument',
                    data={'chat_id': chat_id, 'caption': caption},
                    files={'document': doc},
                    timeout=30,
                )
            print(f'  {GREEN}Via Telegram gesendet.{NC}')
    except Exception:
        pass


# ============================================================
# Hilfsfunktionen
# ============================================================

def load_all_configs() -> list:
    if not os.path.exists(CONFIGS_DIR):
        return []
    files = sorted(f for f in os.listdir(CONFIGS_DIR)
                   if f.startswith('config_') and f.endswith('_mers.json'))
    configs = []
    for fn in files:
        path = os.path.join(CONFIGS_DIR, fn)
        try:
            with open(path) as f:
                cfg = json.load(f)
            cfg['_filename'] = fn
            configs.append(cfg)
        except Exception:
            pass
    return configs


def ask_backtest_config(default_start='2024-01-01'):
    """Interaktive Abfrage von Zeitraum und Kapital."""
    print('\n--- Bitte Konfiguration fuer den Backtest festlegen ---')
    raw = input(f'Startdatum (JJJJ-MM-TT) [Standard: {default_start}]: ').strip()
    start_date = raw if raw else default_start

    raw = input('Enddatum (JJJJ-MM-TT) [Standard: Heute]: ').strip()
    end_date = raw if raw else datetime.now(timezone.utc).strftime('%Y-%m-%d')

    raw = input('Startkapital in USDT eingeben [Standard: 1000]: ').strip()
    try:
        start_capital = float(raw) if raw else 1000.0
    except ValueError:
        start_capital = 1000.0

    print('-' * 50)
    return start_date, end_date, start_capital


def get_exchange_and_risk():
    """Laedt Exchange + Risiko-Config."""
    from mbot.utils.exchange import Exchange
    with open(os.path.join(PROJECT_ROOT, 'secret.json')) as f:
        secrets = json.load(f)
    with open(os.path.join(PROJECT_ROOT, 'settings.json')) as f:
        settings = json.load(f)
    accounts = secrets.get('mbot', [])
    if not accounts:
        print(f'{RED}Keine mbot-Accounts in secret.json.{NC}')
        sys.exit(1)
    exchange    = Exchange(accounts[0])
    risk_config = settings.get('risk', {})
    telegram    = secrets.get('telegram', {})
    return exchange, risk_config, telegram


def run_all_backtests(configs, exchange, risk_config, start_date, end_date, start_capital):
    """Fuehrt Backtest fuer alle Configs durch. Gibt dict {filename: result} zurueck."""
    from mbot.analysis.backtester import load_data, run_backtest
    results = {}
    for cfg in configs:
        fn        = cfg.get('_filename', '?')
        market    = cfg.get('market', {})
        symbol    = market.get('symbol', '?')
        tf        = market.get('timeframe', '?')
        sig_cfg   = cfg.get('signal', {})

        print(f'  Lade: {symbol} ({tf})...')
        df = load_data(exchange, symbol, tf, start_date, end_date)
        if df is None or df.empty:
            print(f'  {RED}Keine Daten. Ueberspringe.{NC}')
            continue

        result = run_backtest(df, sig_cfg, risk_config, start_capital=start_capital,
                               symbol=symbol)
        result['timeframe']  = tf
        result['start_date'] = start_date
        result['end_date']   = end_date
        results[fn] = result
    return results


def print_single_result(result, show_trades=False):
    """Gibt Ergebnis einer einzelnen Strategie aus."""
    symbol   = result.get('symbol', '?')
    tf       = result.get('timeframe', '?')
    trades   = result.get('total_trades', 0)
    win_rate = result.get('win_rate', 0.0)
    pnl_pct  = result.get('total_pnl_pct', 0.0)
    max_dd   = result.get('max_drawdown', 0.0)
    end_cap  = result.get('end_capital', 0.0)
    start_cap= result.get('start_capital', 0.0)

    pnl_c = GREEN if pnl_pct >= 0 else RED
    wr_c  = GREEN if win_rate >= 55 else YELLOW if win_rate >= 45 else RED
    dd_c  = GREEN if max_dd <= 10 else YELLOW if max_dd <= 25 else RED

    print(
        f'  {BOLD}{symbol:<22}{NC} {tf:<6} '
        f'Trades: {trades:>4} | '
        f'WR: {wr_c}{win_rate:>5.1f}%{NC} | '
        f'PnL: {pnl_c}{pnl_pct:>+7.2f}%{NC} ({pnl_c}{end_cap - start_cap:>+8.2f} USDT{NC}) | '
        f'MaxDD: {dd_c}{max_dd:>6.1f}%{NC} | '
        f'End: {end_cap:>8.2f} USDT'
    )

    if show_trades:
        for t in result.get('trades', []):
            side   = t.get('side', '?').upper()
            entry  = t.get('entry_time', '')[:16].replace('T', ' ')
            ep     = t.get('entry_price', 0.0)
            xp     = t.get('exit_price', 0.0)
            res    = t.get('result', '?')
            pnl    = t.get('pnl_pct', 0.0)
            rc     = GREEN if res == 'win' else RED
            print(f'    {entry} | {side:<5} | {ep:>10.4f} → {xp:>10.4f} | '
                  f'{rc}{res:<4}{NC} {pnl:>+6.2f}%')


# ============================================================
# Modus 1: Einzel-Analyse
# ============================================================

def mode_single(target_max_dd):
    print('\n--- mbot Ergebnis-Analyse (Einzel-Modus) ---\n')

    configs = load_all_configs()
    if not configs:
        print(f'{RED}Keine Config-Dateien gefunden in {CONFIGS_DIR}{NC}')
        print(f'{YELLOW}Bitte zuerst run_pipeline.sh ausfuehren.{NC}')
        return

    start_date, end_date, start_capital = ask_backtest_config()
    exchange, risk_config, _ = get_exchange_and_risk()

    print(f'\nZeitraum: {start_date} bis {end_date} | Startkapital: {start_capital} USDT\n')
    results = run_all_backtests(configs, exchange, risk_config, start_date, end_date, start_capital)

    if not results:
        print(f'{RED}Keine Ergebnisse.{NC}')
        return

    print(f'\n{BOLD}{"Symbol":<22} {"TF":<6} {"Trades":>6} {"Win%":>6} {"PnL%":>8} {"USDT":>9} {"MaxDD%":>8} {"Endkap.":>10}{NC}')
    print('-' * 82)

    for fn, r in results.items():
        symbol   = r.get('symbol', '?')
        tf       = r.get('timeframe', '?')
        trades   = r.get('total_trades', 0)
        wr       = r.get('win_rate', 0.0)
        pnl_pct  = r.get('total_pnl_pct', 0.0)
        pnl_usdt = r.get('total_pnl_usdt', 0.0)
        max_dd   = r.get('max_drawdown', 0.0)
        end_cap  = r.get('end_capital', 0.0)

        pnl_c = GREEN if pnl_pct >= 0 else RED
        wr_c  = GREEN if wr >= 55 else YELLOW if wr >= 45 else RED
        dd_c  = GREEN if max_dd <= 10 else YELLOW if max_dd <= 25 else RED

        print(
            f'{symbol:<22} {tf:<6} {trades:>6} '
            f'{wr_c}{wr:>5.1f}%{NC} '
            f'{pnl_c}{pnl_pct:>+7.2f}%{NC} '
            f'{pnl_c}{pnl_usdt:>+8.2f}{NC} '
            f'{dd_c}{max_dd:>7.1f}%{NC} '
            f'{end_cap:>10.2f}'
        )
    print()

    raw = input('  Excel exportieren (alle Trades)? (j/n): ').strip().lower()
    if raw in ('j', 'y', 'ja', 'yes'):
        from mbot.analysis.portfolio_simulator import run_portfolio_simulation
        portfolio = run_portfolio_simulation(results, start_capital)
        _generate_trades_excel(portfolio, start_capital)


# ============================================================
# Modus 2: Manuelle Portfolio-Simulation
# ============================================================

def mode_manual_portfolio(target_max_dd):
    print('\n--- mbot Manuelle Portfolio-Simulation ---\n')

    configs = load_all_configs()
    if not configs:
        print(f'{RED}Keine Configs gefunden. Bitte run_pipeline.sh ausfuehren.{NC}')
        return

    # Verfuegbare Configs auflisten
    print(f'{BOLD}{"="*60}{NC}')
    print('Verfuegbare Strategien:')
    print(f'{BOLD}{"="*60}{NC}')
    for idx, cfg in enumerate(configs, 1):
        market  = cfg.get('market', {})
        symbol  = market.get('symbol', '?')
        tf      = market.get('timeframe', '?')
        pnl     = cfg.get('_meta', {}).get('pnl_pct')
        pnl_str = f'  [{pnl:+.1f}%]' if pnl is not None else ''
        clean   = cfg['_filename'].replace('config_', '').replace('_mers.json', '')
        print(f'{idx:>3}) {clean}{CYAN}{pnl_str}{NC}')
    print(f'{BOLD}{"="*60}{NC}')

    raw = input('\nWaehle Strategien (z.B. "1,3" oder "alle"): ').strip().lower()
    if raw == 'alle' or raw == 'all':
        selected = configs
    else:
        indices = []
        for part in raw.replace(',', ' ').split():
            try:
                indices.append(int(part) - 1)
            except ValueError:
                pass
        selected = [configs[i] for i in indices if 0 <= i < len(configs)]

    if not selected:
        print(f'{RED}Keine gueltigen Strategien ausgewaehlt.{NC}')
        return

    start_date, end_date, start_capital = ask_backtest_config()
    exchange, risk_config, _ = get_exchange_and_risk()

    print(f'\nLade Daten und fuehre Backtests durch...')
    results = run_all_backtests(selected, exchange, risk_config, start_date, end_date, start_capital)

    if not results:
        print(f'{RED}Keine Ergebnisse.{NC}')
        return

    from mbot.analysis.portfolio_simulator import run_portfolio_simulation
    portfolio = run_portfolio_simulation(results, start_capital)
    _print_portfolio_result(portfolio, 'Manuelles Portfolio')

    raw = input('\n  Portfolio-Chart erstellen & via Telegram senden? (j/n): ').strip().lower()
    if raw in ('j', 'y', 'ja', 'yes'):
        from mbot.analysis.interactive_chart import generate_portfolio_chart, _send_charts_via_telegram
        with open(os.path.join(PROJECT_ROOT, 'secret.json')) as f:
            secrets = json.load(f)
        tg        = secrets.get('telegram', {})
        bot_token = tg.get('bot_token', '')
        chat_id   = tg.get('chat_id', '')

        print(f'  Erstelle Portfolio-Chart...')
        path = generate_portfolio_chart(
            results, portfolio,
            start_capital, start_date, end_date,
        )
        if path:
            print(f'{GREEN}  ✓ Portfolio-Chart erstellt: {path}{NC}')
            if bot_token and chat_id:
                _send_charts_via_telegram([path], bot_token, chat_id)
                print(f'{GREEN}  ✓ Via Telegram gesendet.{NC}')

        _generate_trades_excel(portfolio, start_capital)
    else:
        raw = input('  Excel exportieren? (j/n): ').strip().lower()
        if raw in ('j', 'y', 'ja', 'yes'):
            _generate_trades_excel(portfolio, start_capital)


# ============================================================
# Modus 3: Automatische Portfolio-Optimierung
# ============================================================

def mode_auto_portfolio(target_max_dd):
    configs = load_all_configs()
    if not configs:
        print(f'{RED}Keine optimierten Strategien (Configs) gefunden.{NC}')
        return

    start_date, end_date, start_capital = ask_backtest_config()
    exchange, risk_config, telegram = get_exchange_and_risk()

    sep = '─' * 72
    print(f'\n{sep}')
    print(f'  mbot Automatische Portfolio-Optimierung')
    print(f'  Ziel: Maximaler Profit bei maximal {target_max_dd:.1f}% Drawdown.'
          f' | {start_date} → {end_date}')
    print(f'  Modell: Gemeinsamer Kapital-Pool — pro Strategie 1 Trade gleichzeitig moeglich')
    print(f'  Score:  Calmar Ratio (PnL / MaxDD) — balanciert Rendite und Risiko')
    print(f'  Constraint: max. 1 Timeframe pro Coin (bester Calmar wird automatisch gewaehlt)')
    print(f'{sep}\n')

    print(f'  Lade Backtest-Ergebnisse ...', end='', flush=True)
    results = run_all_backtests(configs, exchange, risk_config, start_date, end_date, start_capital)
    with_trades = sum(1 for r in results.values() if r.get('total_trades', 0) > 0)
    print(f' {len(results)} Dateien, {with_trades} mit Trades.')

    if not results:
        print(f'{RED}Keine Ergebnisse.{NC}')
        return

    # Strategien mit negativem PnL ausfiltern
    positive_results = {k: v for k, v in results.items() if v.get('total_pnl_pct', 0.0) > 0}
    if not positive_results:
        print(f'{RED}Keine Strategie mit positivem PnL gefunden.{NC}')
        return
    excluded = len(results) - len(positive_results)
    if excluded:
        print(f'  {excluded} Strategie(n) mit negativem PnL ausgeschlossen.\n')

    print(f'\n  Optimiere Portfolio...\n')
    from mbot.analysis.portfolio_simulator import find_best_portfolio
    best = find_best_portfolio(positive_results, start_capital, target_max_dd, verbose=True)

    if not best:
        print(f'\n{RED}Kein Portfolio gefunden das den Drawdown-Constraint ({target_max_dd}%) erfuellt.{NC}')
        return

    port    = best['portfolio']
    pnl_pct = port.get('total_pnl_pct', 0.0)
    max_dd  = port.get('max_drawdown', 0.0)
    trades  = port.get('total_trades', 0)
    wr      = port.get('win_rate', 0.0)
    end_cap = port.get('end_capital', start_capital)
    pnl_c   = GREEN if pnl_pct >= 0 else RED
    dd_c    = GREEN if max_dd <= 10 else YELLOW if max_dd <= 25 else RED

    print(f'\n{"="*72}')
    print(f'  mbot — Automatische Portfolio-Optimierung')
    print(f'  Ziel: Maximaler Profit bei maximal {target_max_dd:.1f}% Drawdown.')
    print(f'{"="*72}\n')

    n_selected = len(best['selected'])
    print(f'  Optimales Portfolio — {n_selected} Strategie(n)')
    print(f'  Kapital: {start_capital:.0f} USDT  (gemeinsamer Pool)\n')

    # Tabelle pro Strategie
    print(f'  {"Markt":<25} {"TF":<6} {"Trades":>7} {"WR":>8} {"PnL%":>10}  {"MaxDD":>7}')
    print(f'  {"─"*68}')
    for fn in best['selected']:
        r      = results[fn]
        sym    = r.get('symbol', '?')
        tf     = r.get('timeframe', '?')
        tr     = r.get('total_trades', 0)
        w      = r.get('win_rate', 0.0)
        p      = r.get('total_pnl_pct', 0.0)
        dd     = r.get('max_drawdown', 0.0)
        pc     = GREEN if p >= 0 else RED
        dc     = GREEN if dd <= 10 else YELLOW if dd <= 25 else RED
        print(f'  {sym:<25} {tf:<6} {tr:>7}  {w:>5.1f}% {pc}{p:>+9.1f}%{NC}  {dc}{dd:>5.1f}%{NC}')

    print(f'\n  {"─"*68}')
    calmar = pnl_pct / max_dd if max_dd > 0 else pnl_pct
    print(f'  Portfolio gesamt (gemeinsamer Kapital-Pool, pro Strategie 1 gleichzeitiger Trade):')
    print(f'  Trades total:  {trades}')
    print(f'  Win-Rate:      {wr:.1f}%')
    print(f'  PnL:           {pnl_c}{pnl_pct:+.1f}%{NC}')
    print(f'  Calmar Ratio:  {calmar:.2f}')
    print(f'  Final Equity:  {end_cap:.2f} USDT')
    print(f'  Max Drawdown:  {dd_c}{max_dd:.1f}%{NC}')
    print(f'{"="*72}\n')

    # settings.json aktualisieren
    os.makedirs(RESULTS_DIR, exist_ok=True)
    portfolio_data = {
        'selected_strategies': [
            {'symbol': results[fn].get('symbol'), 'timeframe': results[fn].get('timeframe')}
            for fn in best['selected']
        ],
        'pnl_pct':      pnl_pct,
        'max_drawdown': max_dd,
    }
    with open(os.path.join(RESULTS_DIR, 'optimal_portfolio.json'), 'w') as f:
        json.dump(portfolio_data, f, indent=2)

    raw = input('  Sollen die optimalen Ergebnisse in settings.json eingetragen werden? (j/n): ').strip().lower()
    if raw in ('j', 'y', 'ja', 'yes'):
        settings_path = os.path.join(PROJECT_ROOT, 'settings.json')
        with open(settings_path) as f:
            settings = json.load(f)
        strategies = [
            {'symbol': results[fn].get('symbol'), 'timeframe': results[fn].get('timeframe'), 'active': True}
            for fn in best['selected']
        ]
        settings['live_trading_settings']['active_strategies'] = strategies
        with open(settings_path, 'w') as f:
            json.dump(settings, f, indent=4)
        print(f'{GREEN}  ✓ settings.json aktualisiert — {len(strategies)} Strategie(n) eingetragen.{NC}')
    else:
        print(f'{YELLOW}  Keine Aenderungen an settings.json vorgenommen.{NC}')

    # Portfolio-Chart
    raw = input(f'\n  Interaktiver Portfolio-Chart erstellen & via Telegram senden? (j/n): ').strip().lower()
    if raw in ('j', 'y', 'ja', 'yes'):
        from mbot.analysis.interactive_chart import generate_portfolio_chart, _send_charts_via_telegram
        from mbot.analysis.portfolio_simulator import run_portfolio_simulation
        with open(os.path.join(PROJECT_ROOT, 'secret.json')) as f:
            secrets = json.load(f)
        tg        = secrets.get('telegram', {})
        bot_token = tg.get('bot_token', '')
        chat_id   = tg.get('chat_id', '')

        selected_results = {fn: results[fn] for fn in best['selected']}
        # Portfolio neu simulieren um portfolio_capital_after in Trades zu haben
        portfolio_with_trades = run_portfolio_simulation(selected_results, start_capital)

        print(f'  Erstelle Portfolio-Chart...')
        path = generate_portfolio_chart(
            selected_results, portfolio_with_trades,
            start_capital, start_date, end_date,
        )
        if path:
            print(f'{GREEN}  ✓ Portfolio-Chart erstellt: {path}{NC}')
            if bot_token and chat_id:
                _send_charts_via_telegram([path], bot_token, chat_id)
                print(f'{GREEN}  ✓ Via Telegram gesendet.{NC}')

        # Excel Export
        _generate_trades_excel(portfolio_with_trades, start_capital)


def _print_portfolio_result(portfolio, label):
    if not portfolio:
        print(f'{RED}Keine Portfolio-Ergebnisse.{NC}')
        return
    pnl_pct  = portfolio.get('total_pnl_pct', 0.0)
    pnl_usdt = portfolio.get('total_pnl_usdt', 0.0)
    max_dd   = portfolio.get('max_drawdown', 0.0)
    trades   = portfolio.get('total_trades', 0)
    wr       = portfolio.get('win_rate', 0.0)
    end_cap  = portfolio.get('end_capital', 0.0)

    pnl_c = GREEN if pnl_pct >= 0 else RED
    dd_c  = GREEN if max_dd <= 10 else YELLOW if max_dd <= 25 else RED

    print(f'\n{BOLD}{"="*55}{NC}')
    print(f'{BOLD}  {label}{NC}')
    print(f'{BOLD}{"="*55}{NC}')
    print(f'  Trades:      {trades}')
    print(f'  Win-Rate:    {wr:.1f}%')
    print(f'  PnL:         {pnl_c}{pnl_pct:+.2f}%  ({pnl_usdt:+.2f} USDT){NC}')
    print(f'  Max Drawdown:{dd_c}{max_dd:.1f}%{NC}')
    print(f'  Endkapital:  {end_cap:.2f} USDT')
    print(f'{BOLD}{"="*55}{NC}\n')


# ============================================================
# Modus 4: Interaktive Charts
# ============================================================

def mode_interactive_charts():
    from mbot.analysis.interactive_chart import run_interactive_chart
    run_interactive_chart()


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='mbot Ergebnisse anzeigen')
    parser.add_argument('--mode',               type=str, default='1',
                        choices=['1', '2', '3', '4'])
    parser.add_argument('--target_max_drawdown', type=float, default=30.0)
    args = parser.parse_args()

    try:
        if args.mode == '1':
            mode_single(args.target_max_drawdown)
        elif args.mode == '2':
            mode_manual_portfolio(args.target_max_drawdown)
        elif args.mode == '3':
            mode_auto_portfolio(args.target_max_drawdown)
        elif args.mode == '4':
            mode_interactive_charts()
    except KeyboardInterrupt:
        print(f'\n{YELLOW}Abgebrochen.{NC}')
        sys.exit(0)


if __name__ == '__main__':
    main()
